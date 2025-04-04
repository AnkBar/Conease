23import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime
import random
import string

class RequestManagementSystem:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Request Management System")
        self.window.geometry("800x600")
        
        self.setup_excel()
        self.current_user = None
        self.show_login_screen()

    def setup_excel(self):
        try:
            self.wb = openpyxl.load_workbook("request_system.xlsx")
        except:
            self.wb = openpyxl.Workbook()
            ws = self.wb.active
            ws.title = "Requests"
            headers = ["Request ID", "User", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9", "Q10",
                      "Input1", "Input2", "Status", "Approval ID", "Assigned To", "Timestamp", "Acknowledged"]
            ws.append(headers)
            
            self.wb.create_sheet("Users")
            self.wb["Users"].append(["Username", "Password", "Role", "Status"])
            self.wb["Users"].append(["admin", "admin123", "approver", "Approved"])
            self.wb["Users"].append(["user1", "user123", "user", "Approved"])
            self.wb["Users"].append(["user2", "user123", "user", "Approved"])
            self.wb["Users"].append(["user3", "user123", "user", "Approved"])
            
            self.wb.save("request_system.xlsx")

    def generate_id(self, prefix):
        return prefix + ''.join(random.choices(string.digits, k=6))

    def show_login_screen(self):
        for widget in self.window.winfo_children():
            widget.destroy()

        tk.Label(self.window, text="Login", font=("Arial", 16)).pack(pady=20)
        
        tk.Label(self.window, text="Username:").pack()
        self.username_entry = tk.Entry(self.window)
        self.username_entry.pack()
        
        tk.Label(self.window, text="Password:").pack()
        self.password_entry = tk.Entry(self.window, show="*")
        self.password_entry.pack()
        
        tk.Button(self.window, text="Login", command=self.login).pack(pady=10)
        tk.Button(self.window, text="Request New User", command=self.request_new_user).pack()

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        ws = self.wb["Users"]
        for row in ws.rows:
            if row[0].value == username and row[1].value == password and row[3].value == "Approved":
                self.current_user = {"username": username, "role": row[2].value}
                if self.current_user["role"] == "approver":
                    self.show_approver_dashboard()
                else:
                    self.show_user_dashboard()
                return
        messagebox.showerror("Login Failed", "Invalid credentials or user not approved")

    def request_new_user(self):
        new_window = tk.Toplevel(self.window)
        new_window.title("New User Request")
        new_window.geometry("300x200")
        
        tk.Label(new_window, text="Desired Username:").pack(pady=5)
        username = tk.Entry(new_window)
        username.pack()
        
        tk.Label(new_window, text="Password:").pack(pady=5)
        password = tk.Entry(new_window, show="*")
        password.pack()
        
        def submit():
            ws = self.wb["Users"]
            for row in ws.rows:
                if row[0].value == username.get():
                    messagebox.showerror("Error", "User already registered. Please login.")
                    new_window.destroy()
                    return
            ws.append([username.get(), password.get(), "user", "Pending"])
            self.wb.save("request_system.xlsx")
            messagebox.showinfo("Success", "User request submitted for approval")
            new_window.destroy()
        
        tk.Button(new_window, text="Submit Request", command=submit).pack(pady=10)

    def show_user_dashboard(self):
        for widget in self.window.winfo_children():
            widget.destroy()
            
        tk.Label(self.window, text=f"Welcome {self.current_user['username']}", font=("Arial", 16)).pack(pady=20)
        
        tk.Button(self.window, text="Raise New Request", command=self.raise_request).pack(pady=5)
        tk.Button(self.window, text="View/Acknowledge Requests", command=self.view_requests).pack(pady=5)
        tk.Button(self.window, text="Logout", command=self.show_login_screen).pack(pady=5)

    def show_approver_dashboard(self):
        for widget in self.window.winfo_children():
            widget.destroy()
            
        tk.Label(self.window, text=f"Welcome Approver {self.current_user['username']}", font=("Arial", 16)).pack(pady=20)
        
        tk.Label(self.window, text="Assign to:").pack()
        self.assign_combo = ttk.Combobox(self.window, values=["user1", "user2", "user3"])
        self.assign_combo.pack(pady=5)
        
        tk.Button(self.window, text="Review User Requests", command=self.review_user_requests).pack(pady=5)
        tk.Button(self.window, text="Review Form Requests", command=self.review_form_requests).pack(pady=5)
        tk.Button(self.window, text="Logout", command=self.show_login_screen).pack(pady=5)

    def raise_request(self):
        request_window = tk.Toplevel(self.window)
        request_window.title("New Request")
        request_window.geometry("400x600")
        
        answers = {}
        for i in range(10):
            tk.Label(request_window, text=f"Question {i+1}").pack(pady=2)
            combo = ttk.Combobox(request_window, values=["Option 1", "Option 2", "Option 3"])
            combo.pack()
            answers[f"Q{i+1}"] = combo
        
        tk.Label(request_window, text="Input 1:").pack(pady=2)
        input1 = tk.Entry(request_window)
        input1.pack()
        
        tk.Label(request_window, text="Input 2:").pack(pady=2)
        input2 = tk.Entry(request_window)
        input2.pack()
        
        def submit():
            request_id = self.generate_id("R")
            values = [request_id, self.current_user["username"]]
            values.extend([answers[f"Q{i+1}"].get() for i in range(10)])
            values.extend([input1.get(), input2.get(), "Pending", "", "", 
                          datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "No"])
            
            self.wb["Requests"].append(values)
            self.wb.save("request_system.xlsx")
            messagebox.showinfo("Success", f"Request submitted with ID: {request_id}")
            request_window.destroy()
            self.modify_request(request_id)
        
        tk.Button(request_window, text="Submit", command=submit).pack(pady=10)

    def modify_request(self, request_id):
        modify_window = tk.Toplevel(self.window)
        modify_window.title(f"Modify Request {request_id}")
        modify_window.geometry("400x600")
        
        ws = self.wb["Requests"]
        row_data = None
        for row in ws.rows:
            if row[0].value == request_id:
                row_data = row
                break
        
        answers = {}
        for i in range(10):
            tk.Label(modify_window, text=f"Question {i+1}").pack(pady=2)
            combo = ttk.Combobox(modify_window, values=["Option 1", "Option 2", "Option 3"])
            combo.set(row_data[i+2].value)
            combo.pack()
            answers[f"Q{i+1}"] = combo
        
        tk.Label(modify_window, text="Input 1:").pack(pady=2)
        input1 = tk.Entry(modify_window)
        input1.insert(0, row_data[12].value)
        input1.pack()
        
        tk.Label(modify_window, text="Input 2:").pack(pady=2)
        input2 = tk.Entry(modify_window)
        input2.insert(0, row_data[13].value)
        input2.pack()
        
        def save_changes():
            for row in ws.rows:
                if row[0].value == request_id and row[14].value == "Pending":
                    for i in range(10):
                        row[i+2].value = answers[f"Q{i+1}"].get()
                    row[12].value = input1.get()
                    row[13].value = input2.get()
                    break
            self.wb.save("request_system.xlsx")
            messagebox.showinfo("Success", "Request modified successfully")
            modify_window.destroy()
        
        tk.Button(modify_window, text="Save Changes", command=save_changes).pack(pady=10)

    def view_requests(self):
        view_window = tk.Toplevel(self.window)
        view_window.title("My Requests")
        view_window.geometry("900x400")
        
        tree = ttk.Treeview(view_window, columns=("ID", "Status", "Approval ID", "Assigned To", "Acknowledged"), 
                           show="headings")
        tree.heading("ID", text="Request ID")
        tree.heading("Status", text="Status")
        tree.heading("Approval ID", text="Approval ID")
        tree.heading("Assigned To", text="Assigned To")
        tree.heading("Acknowledged", text="Acknowledged")
        tree.pack(fill="both", expand=True)
        
        ws = self.wb["Requests"]
        for row in ws.rows:
            if row[1].value == self.current_user["username"]:
                tree.insert("", "end", values=(row[0].value, row[14].value, row[15].value, 
                                             row[16].value, row[18].value))
        
        def acknowledge():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a request")
                return
                
            request_id = tree.item(selected)["values"][0]
            for row in ws.rows:
                if row[0].value == request_id and row[14].value == "Approved" and row[18].value == "No":
                    row[18].value = "Yes"
                    self.wb.save("request_system.xlsx")
                    messagebox.showinfo("Success", f"Request {request_id} acknowledged and closed")
                    view_window.destroy()
                    self.view_requests()
                    return
            messagebox.showwarning("Warning", "Cannot acknowledge this request")
        
        tk.Button(view_window, text="Acknowledge Selected", command=acknowledge).pack(pady=5)

    def review_user_requests(self):
        review_window = tk.Toplevel(self.window)
        review_window.title("Pending User Requests")
        review_window.geometry("400x300")
        
        tree = ttk.Treeview(review_window, columns=("Username", "Status"), show="headings")
        tree.heading("Username", text="Username")
        tree.heading("Status", text="Status")
        tree.pack(fill="both", expand=True)
        
        ws = self.wb["Users"]
        for row in ws.rows:
            if row[3].value == "Pending":
                tree.insert("", "end", values=(row[0].value, row[3].value))
        
        def approve():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a user")
                return
            username = tree.item(selected)["values"][0]
            for row in ws.rows:
                if row[0].value == username:
                    row[3].value = "Approved"
                    break
            self.wb.save("request_system.xlsx")
            messagebox.showinfo("Success", f"User {username} approved")
            review_window.destroy()
            self.review_user_requests()
        
        def reject():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a user")
                return
            username = tree.item(selected)["values"][0]
            for row in ws.rows:
                if row[0].value == username:
                    ws.delete_rows(row[0].row, 1)
                    break
            self.wb.save("request_system.xlsx")
            messagebox.showinfo("Success", f"User {username} rejected")
            review_window.destroy()
            self.review_user_requests()
        
        tk.Button(review_window, text="Approve", command=approve).pack(side="left", padx=5, pady=5)
        tk.Button(review_window, text="Reject", command=reject).pack(side="left", padx=5, pady=5)

    def review_form_requests(self):
        review_window = tk.Toplevel(self.window)
        review_window.title("Form Requests")
        review_window.geometry("1000x500")
        
        tree = ttk.Treeview(review_window, columns=("ID", "User", "Status", "Approval ID", "Assigned To"), 
                           show="headings")
        tree.heading("ID", text="Request ID")
        tree.heading("User", text="User")
        tree.heading("Status", text="Status")
        tree.heading("Approval ID", text="Approval ID")
        tree.heading("Assigned To", text="Assigned To")
        tree.pack(fill="both", expand=True)
        
        ws = self.wb["Requests"]
        for row in ws.rows:
            if row[0].value != "Request ID":  # Skip header
                tree.insert("", "end", values=(row[0].value, row[1].value, row[14].value, 
                                             row[15].value, row[16].value))
        
        def approve():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a request")
                return
            if not self.assign_combo.get():
                messagebox.showwarning("Warning", "Please assign a user first")
                return
            request_id = tree.item(selected)["values"][0]
            approval_id = self.generate_id("A")
            for row in ws.rows:
                if row[0].value == request_id and row[14].value == "Pending":
                    row[14].value = "Approved"
                    row[15].value = approval_id
                    row[16].value = self.assign_combo.get()
                    break
            self.wb.save("request_system.xlsx")
            messagebox.showinfo("Success", f"Request {request_id} approved with ID: {approval_id}")
            review_window.destroy()
            self.review_form_requests()
        
        def reject():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a request")
                return
            if not self.assign_combo.get():
                messagebox.showwarning("Warning", "Please assign a user first")
                return
            request_id = tree.item(selected)["values"][0]
            for row in ws.rows:
                if row[0].value == request_id and row[14].value == "Pending":
                    row[14].value = "Rejected"
                    row[16].value = self.assign_combo.get()
                    break
            self.wb.save("request_system.xlsx")
            messagebox.showinfo("Success", f"Request {request_id} rejected")
            review_window.destroy()
            self.review_form_requests()
        
        def modify():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select a request")
                return
            request_id = tree.item(selected)["values"][0]
            self.modify_request(request_id)
            review_window.destroy()
        
        tk.Button(review_window, text="Approve", command=approve).pack(side="left", padx=5, pady=5)
        tk.Button(review_window, text="Reject", command=reject).pack(side="left", padx=5, pady=5)
        tk.Button(review_window, text="Modify", command=modify).pack(side="left", padx=5, pady=5)

    def run(self):
        self.window.mainloop()

if __name__ == "__main__":
    app = RequestManagementSystem()
    app.run()
    from flask import Flask
